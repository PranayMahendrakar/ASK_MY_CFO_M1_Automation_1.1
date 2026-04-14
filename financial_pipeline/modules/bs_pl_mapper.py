#!/usr/bin/env python3
"""
BS/PL Financial Statement → Template Mapper (v2)
=================================================
Converts extracted BS/PL Excel files into a standardized template using
OpenAI GPT-4o for intelligent label mapping.

Key accuracy features:
  - Residual-based F72 computation (P&L guaranteed to balance)
  - BS section-total cross-check with auto-retry
  - Combined single-file output (Standalone + Consolidated × Current + Prior)
  - Multi-row label merging with section-header detection

Usage:
  Single file:   python bs_pl_mapper.py --input file.xlsx --template tmpl.xlsx --api-key sk-... --output out/
  Batch folder:  python bs_pl_mapper.py --input folder/    --template tmpl.xlsx --api-key sk-... --output out/
"""

import os, sys, json, re, argparse
from pathlib import Path
from typing import Optional
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from openai import OpenAI
except ImportError:
    os.system(f"{sys.executable} -m pip install openai --break-system-packages -q")
    from openai import OpenAI

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PREPROCESSING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

NON_MERGE_LABELS = {
    "income", "expenses", "assets", "liabilities", "equity",
    "financial assets", "non-current assets", "current assets",
    "non-current liabilities", "current liabilities",
    "non- current assets", "non- current liabilities",
    "equity and liabilities", "shareholder's fund", "shareholders fund",
    "trade payables", "tax expense", "tax expenses",
    "other comprehensive income", "exceptional items",
    "items that will not be reclassified to profit or loss:",
    "items that will be reclassified to profit or loss:",
    "property, plant & equipment", "revenue",
    "borrowings", "provisions", "investments", "loans",
    "intangible assets", "financial liabilities",
    "non-controlling interest", "non controlling interest",
}

def clean_value(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("(", "-").replace(")", "")
    if s in ["-", "–", "—", "", "nil", "Nil", "NIL"]:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def is_section_header(label):
    return label.lower().strip().rstrip(":") in NON_MERGE_LABELS

def merge_multirow(items):
    merged, skip = [], False
    for i, item in enumerate(items):
        if skip:
            skip = False
            continue
        if is_section_header(item["label"]):
            merged.append(item)
            continue
        if i + 1 < len(items):
            nxt = items[i + 1]
            cur_has = item["cur"] != 0 or item["pri"] != 0
            nxt_has = nxt["cur"] != 0 or nxt["pri"] != 0
            if not cur_has and nxt_has:
                cl = item["label"].lower().strip()
                if cl.endswith(("of", "the", "than", "to", "and", "other", "micro", "small")):
                    merged.append({"label": item["label"]+" "+nxt["label"], "cur": nxt["cur"], "pri": nxt["pri"]})
                    skip = True
                    continue
                nl = nxt["label"].lower().strip()
                if nl.startswith(("enterprises", "profit or loss", "and small")):
                    merged.append({"label": item["label"]+" "+nxt["label"], "cur": nxt["cur"], "pri": nxt["pri"]})
                    skip = True
                    continue
        merged.append(item)
    return merged

def extract_items(filepath, sheet_name):
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    items, started = [], False
    for _, row in df.iterrows():
        label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if "particulars" in label.lower() and not started:
            started = True
            continue
        if not started or not label or label.lower() == "nan":
            continue
        items.append({
            "label": label,
            "cur": clean_value(row.iloc[2] if len(row) > 2 else None),
            "pri": clean_value(row.iloc[3] if len(row) > 3 else None),
        })
    return merge_multirow(items)

def classify_sheets(sheet_names):
    result = {}
    for name in sheet_names:
        nl = name.lower()
        is_bs = any(k in nl for k in ["balance sheet", "balance_sheet", "bs"])
        is_pl = any(k in nl for k in ["p&l", "p & l", "pl", "profit", "loss", "income"])
        is_con = "consol" in nl
        is_sa = "standalone" in nl or "stand" in nl
        stype = "bs" if is_bs else ("pl" if is_pl else None)
        if not stype:
            continue
        variant = "consolidated" if is_con else ("standalone" if is_sa else "default")
        result[name] = {"type": stype, "variant": variant}
    return result

def process_source(filepath):
    xls = pd.ExcelFile(filepath)
    classified = classify_sheets(xls.sheet_names)
    variants = {}
    for sheet_name, info in classified.items():
        items = extract_items(filepath, sheet_name)
        v = info["variant"]
        if v not in variants:
            variants[v] = {}
        variants[v][info["type"]] = items
    if list(variants.keys()) == ["default"]:
        variants["standalone"] = variants.pop("default")
    return variants


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LLM MAPPING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

BS_PROMPT = """You are an expert Indian financial analyst (Ind AS + Old GAAP).
Map these Balance Sheet line items into the template cells below.

SOURCE (label: current_year_value | prior_year_value):
{items}

TEMPLATE CELLS:
NETWORTH:
  F5  = Share Capital
  F6  = Retained Earnings (0 if not separately disclosed)
  F7  = General Reserves & Surplus (Old GAAP "Reserves and surplus" goes here)
  F8  = Other Equity (Ind AS "Other Equity" goes here)
  F9  = spare (Non-Controlling Interest if present)

CURRENT LIABILITIES:
  F14 = Accounts Payable (SUM all Trade Payables sub-items: MSME + non-MSME)
  F15 = Provisions (current)
  F16 = Short term borrowings
  F17 = Other Current Liabilities
  F18 = Other Financial Liabilities
  F19 = Lease Liabilities (current) — if not present, 0

NON-CURRENT LIABILITIES:
  F24 = Long Term borrowings
  F25 = Provision (non-current)
  F26 = Others = Lease Liabilities(NC) + Deferred Tax Liabilities + Other Non-Current Liabilities (SUM all)
  F27 = Other Financial Liabilities (non-current)
  F28 = spare

CURRENT ASSETS:
  F35 = Bank Balance ("Other Bank Balances"; if only "Cash and Bank Balances" exists as combined, put 0 here)
  F36 = Cash & Cash Equivalence (if combined "Cash and Bank Balances", put full amount here)
  F37 = Inventory / Inventories (0 for IT/service companies)
  F38 = Investments (current)
  F39 = Loans (current) / Short Term Loan and Advances
  F40 = Accounts Receivable = Trade Receivables
  F41 = Other Current Assets + Current Tax Assets (SUM both)
  F42 = Other Financial Assets (current)
  F43 = spare

NON-CURRENT ASSETS:
  F48 = Fixed Assets = PPE + Intangible Assets + Right of Use Assets + Goodwill + Software (SUM all EXCEPT CWIP)
  F49 = Investments (non-current)
  F50 = Loans (non-current) / Long-term loans & advances
  F51 = Capital Work-in-Progress + Intangible Assets under Development (SUM)
  F52 = Other Non-Current Assets
  F53 = Other Financial Assets (non-current)
  F54 = Deferred Tax Assets
  F55 = spare

CRITICAL RULES:
- Every source value MUST appear in exactly one cell. Double-counting = wrong. Missing items = wrong.
- The test: SUM(F5:F9) + SUM(F14:F19) + SUM(F24:F28) MUST equal SUM(F35:F43) + SUM(F48:F55).
- If the source has "Total Assets" and "Total Equity and Liabilities", your mapped totals must match those.

Return ONLY valid JSON (no markdown fences):
{{"current_year":{{"F5":0,...}},"prior_year":{{"F5":0,...}},"notes":"brief explanation"}}"""

PL_PROMPT = """You are an expert Indian financial analyst.
Map these P&L line items to template cells. IMPORTANT: I will back-calculate F72 myself, so DO NOT include F72.

SOURCE (label: current_year_value | prior_year_value):
{items}

TEMPLATE CELLS:
  F66 = Revenue from Operations ONLY (exclude Other Income)
  F67 = Cost of Goods Sold. Rules:
        - Manufacturing: "Cost of Materials Consumed" (+ "Changes in Inventories" if present)
        - Trading: "Purchase of traded goods" + "Changes in inventories of traded goods"
        - Services/IT: "Operating Expenses" if that's the main cost line
        - If no clear COGS exists, put 0
  F69 = Employee Benefits Expense
  F70 = Finance Costs / Interest
  F71 = Depreciation and Amortization Expense
  F73 = Total Tax Expense (SUM of Current Tax + Deferred Tax + Tax of Earlier Years)

ALSO EXTRACT:
  other_income = Other Income value
  net_profit = The final "Profit after tax" / "Profit for the year" / "Net Profit" figure
  exceptional_items = Any exceptional income (positive) or expense (negative). 0 if none.

I will compute F72 = F66 - F67 - F69 - F70 - F71 - F73 - net_profit
This guarantees the P&L balances perfectly.

Return ONLY valid JSON (no markdown fences):
{{"current_year":{{"F66":0,"F67":0,"F69":0,"F70":0,"F71":0,"F73":0,"other_income":0,"net_profit":0,"exceptional_items":0}},"prior_year":{{"F66":0,"F67":0,"F69":0,"F70":0,"F71":0,"F73":0,"other_income":0,"net_profit":0,"exceptional_items":0}},"notes":"brief"}}"""


def format_items_for_prompt(items):
    lines = []
    for it in items:
        if it["cur"] != 0 or it["pri"] != 0:
            lines.append(f'  "{it["label"]}": {it["cur"]} | {it["pri"]}')
    return "\n".join(lines)

def call_gpt(client, prompt, context, max_retries=2):
    for attempt in range(max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "Expert Indian financial analyst. Return ONLY valid JSON. No markdown fences. No explanations outside JSON."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.0,
                max_tokens=3000,
            )
            text = resp.choices[0].message.content.strip()
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
            return json.loads(text)
        except (json.JSONDecodeError, Exception) as e:
            if attempt < max_retries:
                print(f"    [Retry {attempt+1}] {context}: {e}")
            else:
                print(f"    [FAILED] {context}: {e}")
                return None

def map_bs(client, items):
    prompt = BS_PROMPT.replace("{items}", format_items_for_prompt(items))
    return call_gpt(client, prompt, "BS")

def map_pl(client, items):
    prompt = PL_PROMPT.replace("{items}", format_items_for_prompt(items))
    return call_gpt(client, prompt, "PL")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# VALIDATION & RESIDUAL COMPUTATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def validate_bs(mapping, year):
    m = mapping[year]
    nw  = sum(float(m.get(f"F{i}", 0)) for i in range(5, 10))
    cl  = sum(float(m.get(f"F{i}", 0)) for i in range(14, 20))
    ncl = sum(float(m.get(f"F{i}", 0)) for i in range(24, 29))
    ca  = sum(float(m.get(f"F{i}", 0)) for i in range(35, 44))
    nca = sum(float(m.get(f"F{i}", 0)) for i in range(48, 56))
    total_liab = round(nw + cl + ncl, 2)
    total_assets = round(ca + nca, 2)
    diff = round(total_assets - total_liab, 2)
    return {
        "balanced": abs(diff) < 0.5,
        "diff": diff,
        "total_assets": total_assets,
        "total_liabilities": total_liab,
        "networth": round(nw, 2),
        "current_liab": round(cl, 2),
        "noncurrent_liab": round(ncl, 2),
        "current_assets": round(ca, 2),
        "noncurrent_assets": round(nca, 2),
    }

def compute_f72_residual(pl_mapping, year):
    """Back-calculate F72 so P&L is guaranteed to balance."""
    m = pl_mapping[year]
    revenue = float(m["F66"])
    cogs = float(m["F67"])
    emp = float(m["F69"])
    interest = float(m["F70"])
    dep = float(m["F71"])
    tax = float(m["F73"])
    net_profit = float(m["net_profit"])
    # Net Profit = Revenue - COGS - Emp - Interest - Dep - F72 - Tax
    # Therefore: F72 = Revenue - COGS - Emp - Interest - Dep - Tax - Net Profit
    f72 = revenue - cogs - emp - interest - dep - tax - net_profit
    return round(f72, 2)

def retry_bs_with_error(client, items, mapping, year, error_info):
    """Re-prompt LLM with specific error for self-correction."""
    items_text = format_items_for_prompt(items)
    retry_prompt = f"""{BS_PROMPT.replace("{items}", items_text)}

YOUR PREVIOUS ATTEMPT HAD THIS ERROR:
{error_info}

Fix it. Every source value must land in exactly one cell. Check your arithmetic."""
    result = call_gpt(client, retry_prompt, f"BS_retry_{year}")
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL OUTPUT (SINGLE COMBINED FILE)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ROWS = [
    (3, "Balance Sheet Particulars", "header", None),
    (4, "Networth", "section", None),
    (5, "Share Capital", "data", "F5"),
    (6, "Retained Earnings", "data", "F6"),
    (7, "General Reserves & Surplus", "data", "F7"),
    (8, "Other Equity", "data", "F8"),
    (9, "(NCI / spare)", "data", "F9"),
    (10, "Total Networth", "total", "SUM({c}5:{c}9)"),
    (11, "", "blank", None),
    (12, "Current Liabilities", "section", None),
    (14, "Accounts Payable", "data", "F14"),
    (15, "Provisions", "data", "F15"),
    (16, "Short term borrowings", "data", "F16"),
    (17, "Other Current Liabilities", "data", "F17"),
    (18, "Other Financial Liabilities", "data", "F18"),
    (19, "Lease Liabilities (Current)", "data", "F19"),
    (20, "Total Current Liabilities", "total", "SUM({c}14:{c}19)"),
    (21, "", "blank", None),
    (22, "Non - Current Liabilities", "section", None),
    (24, "Long Term borrowings", "data", "F24"),
    (25, "Provision", "data", "F25"),
    (26, "Others (Lease NC + DTL + Other NCL)", "data", "F26"),
    (27, "Other Financial Liabilities", "data", "F27"),
    (28, "(spare)", "data", "F28"),
    (29, "Total Non-Current Liabilities", "total", "SUM({c}24:{c}28)"),
    (30, "", "blank", None),
    (31, "Total Liabilities", "total", "{c}29+{c}20+{c}10"),
    (32, "", "blank", None),
    (33, "Current Assets", "section", None),
    (35, "Bank Balance", "data", "F35"),
    (36, "Cash & Cash Equivalence", "data", "F36"),
    (37, "Inventory", "data", "F37"),
    (38, "Investments", "data", "F38"),
    (39, "Loans", "data", "F39"),
    (40, "Accounts Receivable", "data", "F40"),
    (41, "Other Current Assets", "data", "F41"),
    (42, "Other Financial Assets", "data", "F42"),
    (43, "(spare)", "data", "F43"),
    (44, "Total Current Assets", "total", "SUM({c}35:{c}43)"),
    (45, "", "blank", None),
    (46, "Non - Current Assets", "section", None),
    (48, "Fixed Assets (PPE+Intangible+ROU+Goodwill)", "data", "F48"),
    (49, "Investments", "data", "F49"),
    (50, "Loans", "data", "F50"),
    (51, "Capital Work-in-Progress", "data", "F51"),
    (52, "Other Non-Current Assets", "data", "F52"),
    (53, "Other Financial Assets", "data", "F53"),
    (54, "Deferred Tax Assets", "data", "F54"),
    (55, "(spare)", "data", "F55"),
    (56, "Total Non-Current Assets", "total", "SUM({c}48:{c}55)"),
    (57, "", "blank", None),
    (59, "Total Assets", "total", "{c}56+{c}44"),
    (60, "", "blank", None),
    (61, "DIFFERENCE (must be 0)", "diff", "{c}59-{c}31"),
    (62, "", "blank", None),
    (65, "P&L Statement Particulars", "header", None),
    (66, "Revenue", "data", "F66"),
    (67, "Cost of goods sold", "data", "F67"),
    (68, "Gross profit", "total", "{c}66-{c}67"),
    (69, "Employee benefits expense", "data", "F69"),
    (70, "Interest", "data", "F70"),
    (71, "Depreciation", "data", "F71"),
    (72, "Other expenses less other income", "data", "F72"),
    (73, "Taxes", "data", "F73"),
    (74, "Net Profit", "total", "{c}68-SUM({c}69:{c}73)"),
]

# Styles
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
SEC_FONT = Font(name="Arial", bold=False, size=9, color="4F46E5")
TOT_FILL = PatternFill("solid", fgColor="D9D9D9")
TOT_FONT = Font(name="Arial", bold=True, size=9, color="333333")
DIFF_FILL = PatternFill("solid", fgColor="4F46E5")
DIFF_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
DATA_FONT = Font(name="Arial", size=10, color="0000FF")
FORMULA_FONT = Font(name="Arial", bold=True, size=10, color="000000")
NUM_FMT = '#,##0.00;(#,##0.00);"-"'
THIN_BORDER = Border(bottom=Side(style="thin", color="EEEEEE"))


def build_report(all_data, output_path, company_name=""):
    """Build a single combined Excel report.
    
    all_data = {
        "standalone": {"current": {cells}, "prior": {cells}} or None,
        "consolidated": {"current": {cells}, "prior": {cells}} or None,
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # Determine which columns we need
    columns = []
    col_labels = {}
    col_idx = 5  # start at column E
    for variant in ["standalone", "consolidated"]:
        if variant not in all_data or not all_data[variant]:
            continue
        for year in ["current", "prior"]:
            if year not in all_data[variant]:
                continue
            cl = get_col_letter(col_idx)
            columns.append((cl, col_idx, variant, year))
            v_label = "Standalone" if variant == "standalone" else "Consolidated"
            y_label = "Current Year" if year == "current" else "Prior Year"
            col_labels[cl] = f"{v_label}\n{y_label}"
            col_idx += 1

    if not columns:
        print("  [ERROR] No data to output!")
        return None

    # Column widths
    ws.column_dimensions["D"].width = 44
    for cl, _, _, _ in columns:
        ws.column_dimensions[cl].width = 18

    # Row 1: Company name
    ws["D1"] = company_name or "Financial Statement Report"
    ws["D1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")

    # Row 2: Column headers
    ws["D2"] = "Particulars"
    ws["D2"].font = HDR_FONT
    ws["D2"].fill = HDR_FILL
    for cl, ci, _, _ in columns:
        c = ws.cell(row=2, column=ci, value=col_labels[cl])
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 36

    # Build rows
    for row_num, label, rtype, key in ROWS:
        d = ws.cell(row=row_num, column=4, value=label)

        if rtype == "header":
            d.font = HDR_FONT
            d.fill = HDR_FILL
            d.alignment = Alignment(horizontal="center")
            for _, ci, _, _ in columns:
                ws.cell(row=row_num, column=ci).fill = HDR_FILL
        elif rtype == "section":
            d.font = SEC_FONT
        elif rtype == "total":
            d.font = TOT_FONT
            d.fill = TOT_FILL
        elif rtype == "diff":
            d.font = DIFF_FONT
            d.fill = DIFF_FILL

        for cl, ci, variant, year in columns:
            cell = ws.cell(row=row_num, column=ci)
            if rtype == "data" and key:
                val = all_data[variant][year].get(key, 0)
                cell.value = float(val) if val else 0.0
                cell.font = DATA_FONT
                cell.number_format = NUM_FMT
                cell.border = THIN_BORDER
            elif rtype == "total" and key:
                cell.value = "=" + key.replace("{c}", cl)
                cell.font = FORMULA_FONT
                cell.fill = TOT_FILL
                cell.number_format = NUM_FMT
            elif rtype == "diff" and key:
                cell.value = "=" + key.replace("{c}", cl)
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                cell.fill = DIFF_FILL
                cell.number_format = NUM_FMT

    ws.freeze_panes = "E3"
    wb.save(output_path)
    return output_path

def get_col_letter(idx):
    return chr(ord("A") + idx - 1)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN PIPELINE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def process_file(input_file, template_file, api_key, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    fname = Path(input_file).stem
    print(f"\n{'━'*60}")
    print(f"  Processing: {fname}")
    print(f"{'━'*60}")

    # Stage 1: Preprocess
    print("\n  [1/3] Preprocessing...")
    variants = process_source(input_file)
    for v, sheets in variants.items():
        for t, items in sheets.items():
            real = [i for i in items if i["cur"] != 0 or i["pri"] != 0]
            print(f"    {v}/{t}: {len(real)} valued items")

    # Stage 2: LLM Mapping
    print("\n  [2/3] Mapping via GPT-4o...")
    client = OpenAI(api_key=api_key)
    all_data = {}

    for variant, sheets in variants.items():
        print(f"\n    ── {variant.upper()} ──")
        all_data[variant] = {}

        # Map BS
        bs_mapping = None
        if "bs" in sheets:
            print("    BS: calling GPT-4o...")
            bs_mapping = map_bs(client, sheets["bs"])
            if bs_mapping:
                # Validate both years, retry if needed
                for year in ["current_year", "prior_year"]:
                    if year not in bs_mapping:
                        continue
                    v = validate_bs(bs_mapping, year)
                    yl = year.replace("_year", "")
                    if v["balanced"]:
                        print(f"    BS {yl}: ✓ balanced (Assets={v['total_assets']}, Liab={v['total_liabilities']})")
                    else:
                        print(f"    BS {yl}: ✗ diff={v['diff']} — retrying...")
                        error_info = (
                            f"DIFF = {v['diff']}. "
                            f"Assets side: CA={v['current_assets']}, NCA={v['noncurrent_assets']}, Total={v['total_assets']}. "
                            f"Liab side: NW={v['networth']}, CL={v['current_liab']}, NCL={v['noncurrent_liab']}, Total={v['total_liabilities']}."
                        )
                        retry = retry_bs_with_error(client, sheets["bs"], bs_mapping, year, error_info)
                        if retry and year in retry:
                            v2 = validate_bs(retry, year)
                            if v2["balanced"]:
                                bs_mapping[year] = retry[year]
                                print(f"    BS {yl}: ✓ retry succeeded!")
                            else:
                                print(f"    BS {yl}: ✗ still off by {v2['diff']} — FLAGGED FOR REVIEW")
            else:
                print("    BS: ✗ mapping failed!")

        # Map PL
        pl_mapping = None
        if "pl" in sheets:
            print("    PL: calling GPT-4o...")
            pl_mapping = map_pl(client, sheets["pl"])
            if pl_mapping:
                print(f"    PL: ✓ mapped. Notes: {pl_mapping.get('notes', '')[:80]}")
            else:
                print("    PL: ✗ mapping failed!")

        # Assemble cell values for each year
        for year_key, year_label in [("current_year", "current"), ("prior_year", "prior")]:
            cells = {}
            if bs_mapping and year_key in bs_mapping:
                for k, v in bs_mapping[year_key].items():
                    if k.startswith("F") and k[1:].isdigit():
                        cells[k] = float(v)
            if pl_mapping and year_key in pl_mapping:
                m = pl_mapping[year_key]
                for k in ["F66", "F67", "F69", "F70", "F71", "F73"]:
                    if k in m:
                        cells[k] = float(m[k])
                # BACK-CALCULATE F72 (the killer accuracy feature)
                f72 = compute_f72_residual(pl_mapping, year_key)
                cells["F72"] = f72
                expected_np = float(m.get("net_profit", 0))
                print(f"    PL {year_label}: F72={f72:.2f} (residual), Net Profit={expected_np:.2f}")

            all_data[variant][year_label] = cells

    # Stage 3: Build combined report
    print("\n  [3/3] Building combined report...")
    output_path = os.path.join(output_dir, f"{fname}_Report.xlsx")
    build_report(all_data, output_path, company_name=fname)

    # Recalculate formulas if possible
    recalc = Path(__file__).parent / "recalc.py"
    if not recalc.exists():
        # Try standard location
        recalc = Path("/mnt/skills/public/xlsx/scripts/recalc.py")
    if recalc.exists():
        os.system(f"python {recalc} {output_path}")

    # Final verification
    try:
        wb = load_workbook(output_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        print(f"\n  ── FINAL RESULTS ──")
        for ci in range(5, 5 + len([v for v in all_data.values() if v])*2):
            cl = get_col_letter(ci)
            diff_val = ws.cell(row=61, column=ci).value
            np_val = ws.cell(row=74, column=ci).value
            ta_val = ws.cell(row=59, column=ci).value
            hdr = ws.cell(row=2, column=ci).value or cl
            hdr_short = hdr.replace("\n", " ")
            if ta_val is not None:
                diff_str = f"{diff_val}" if diff_val else "0"
                print(f"    {hdr_short}: Assets={ta_val}, DIFF={diff_str}, NP={np_val}")
    except Exception:
        pass

    print(f"\n  ✓ Output: {output_path}")
    return output_path


def process_batch(input_dir, template_file, api_key, output_dir):
    files = sorted(Path(input_dir).glob("*_extracted*.xlsx"))
    if not files:
        files = sorted(Path(input_dir).glob("*.xlsx"))
        files = [f for f in files if "template" not in f.name.lower() and "output" not in f.name.lower() and "report" not in f.name.lower()]
    print(f"Found {len(files)} file(s) to process.\n")
    for f in files:
        try:
            process_file(str(f), template_file, api_key, output_dir)
        except Exception as e:
            print(f"  ERROR on {f.name}: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BS/PL → Template Mapper v2")
    parser.add_argument("--input", required=True, help="Input .xlsx file or folder for batch")
    parser.add_argument("--template", required=True, help="Template .xlsx file")
    parser.add_argument("--api-key", required=True, help="OpenAI API key")
    parser.add_argument("--output", default="./output", help="Output directory")
    args = parser.parse_args()

    if os.path.isdir(args.input):
        process_batch(args.input, args.template, args.api_key, args.output)
    else:
        process_file(args.input, args.template, args.api_key, args.output)