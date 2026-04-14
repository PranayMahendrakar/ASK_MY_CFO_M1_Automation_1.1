"""
PDF Financial Table Extractor — Final Version
Uses header-anchored column detection for accurate parsing across different layouts.
"""
import re, os
from pathlib import Path
from collections import defaultdict
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

STOP_PHRASES = [
    'summary of', 'accounting polic', 'accompanying notes', 'as per our report',
    'chartered accountant', 'firm registration', 'icai firm', 'membership no',
    'place:', 'date:', 'sd/-', 'din:', 'pan:', 'partner', ' director',
    'company secretary', 'chief financial', 'whole time',
    'on behalf of', 'for and on behalf', 'cin:', 'material accounting'
]

STMT_KEYWORDS = {
    'consolidated_pl': ['CONSOLIDATED', 'PROFIT'],
    'standalone_pl': ['STANDALONE', 'PROFIT'],
    'consolidated_bs': ['CONSOLIDATED', 'BALANCE SHEET'],
    'standalone_bs': ['STANDALONE', 'BALANCE SHEET'],
    'generic_bs': ['BALANCE SHEET'],
    'generic_pl': ['PROFIT'],
}

STMT_NAMES = {
    'consolidated_pl': 'Consolidated P&L',
    'standalone_pl': 'Standalone P&L',
    'consolidated_bs': 'Consolidated Balance Sheet',
    'standalone_bs': 'Standalone Balance Sheet',
    'generic_bs': 'Balance Sheet',
    'generic_pl': 'P&L',
}


def classify_page(text):
    upper = text[:500].upper()
    for key, keywords in STMT_KEYWORDS.items():
        if all(k in upper for k in keywords):
            return key
    return None


def find_column_boundaries(words, page_width):
    """
    Anchor columns from header words: 'Note'/'Notes' and year date columns.
    """
    note_word = None
    year_words = []
    
    for w in words:
        if w['top'] > 200:
            break
        txt = w['text'].strip()
        low = txt.lower()
        if low in ('note', 'notes'):
            note_word = w
        # Only consider year words in the right portion (column headers, not titles)
        if re.match(r'^20\d{2}$', txt) and w['x0'] > page_width * 0.55:
            year_words.append(w)
    
    cy_center = None
    py_center = None
    
    if len(year_words) >= 2:
        year_words.sort(key=lambda w: w['x0'])
        cy_center = (year_words[0]['x0'] + year_words[0]['x1']) / 2
        py_center = (year_words[-1]['x0'] + year_words[-1]['x1']) / 2
    
    if cy_center is None or py_center is None:
        return _fallback_column_detection(words, page_width, note_word)
    
    mid_cy_py = (cy_center + py_center) / 2
    
    # Find actual note number positions and first CY value positions from data
    note_data_rights = []
    cy_data_lefts = []
    for w in words:
        if w['top'] < 100:
            continue
        txt = w['text'].strip()
        x0 = w['x0']
        # Note numbers: 1-2 digit integers in the note column area
        if (re.match(r'^\d{1,2}$', txt) and int(txt) <= 40
                and note_word and abs(x0 - note_word['x0']) < 30):
            note_data_rights.append(w['x1'])
        # CY values: numeric values left of midpoint
        elif (re.match(r'^[\d,(.)]+$', txt.replace(' ', '')) 
              and x0 > (note_word['x1'] + 10 if note_word else page_width * 0.5)
              and (x0 + w['x1'])/2 < mid_cy_py):
            cy_data_lefts.append(x0)
    
    if note_word:
        note_x0 = note_word['x0']
        # Use actual note data extent if available, otherwise header + small margin
        if note_data_rights:
            note_right = max(note_data_rights) + 8
        else:
            note_right = note_word['x1'] + 15
    else:
        note_x0 = cy_center - 60
        note_right = cy_center - 40
    
    # CY left: use actual leftmost CY value, or note_right + gap
    if cy_data_lefts:
        cy_left = min(cy_data_lefts) - 5
    else:
        cy_left = note_right + 5
    
    # Find actual value boundaries from data
    # CY values should end near cy_center + some offset
    cy_val_rights = []
    py_val_rights = []
    for w in words:
        if w['top'] < 100:
            continue
        txt = w['text'].strip().replace(',', '').replace('(', '').replace(')', '').replace(' ', '')
        if not re.match(r'^[\d.]+$', txt) or txt.count('.') > 1:
            continue
        x_mid = (w['x0'] + w['x1']) / 2
        if x_mid < mid_cy_py and w['x0'] > cy_left - 10:
            cy_val_rights.append(w['x1'])
        elif x_mid > mid_cy_py:
            py_val_rights.append(w['x1'])
    
    return {
        'particulars_right': note_x0 - 5 if note_word else cy_left - 30,
        'note_left': note_x0 - 5 if note_word else cy_left - 30,
        'note_right': note_right,
        'cy_left': cy_left,
        'cy_right': mid_cy_py,
        'py_left': mid_cy_py,
        'py_right': max(py_val_rights) + 20 if py_val_rights else page_width,
    }


def _fallback_column_detection(words, page_width, note_word):
    """Fallback when year headers aren't found."""
    num_positions = []
    for w in words:
        if w['top'] < 100:
            continue
        txt = w['text'].strip().replace(',', '').replace('(', '').replace(')', '').replace(' ', '')
        if re.match(r'^[\d,]+\.?\d{0,2}$', txt) and txt.count('.') <= 1:
            try:
                fval = float(txt.replace(',', ''))
            except ValueError:
                continue
            if fval > 0.001:
                num_positions.append((w['x0'], w['x1']))
    
    if len(num_positions) < 4:
        return None
    
    # Cluster right edges
    rights = sorted(set(round(p[1], 0) for p in num_positions))
    clusters = []
    cur = [rights[-1]]
    for x in reversed(rights[:-1]):
        if cur[-1] - x < 20:
            cur.append(x)
        else:
            clusters.append(cur)
            cur = [x]
            if len(clusters) >= 3:
                break
    clusters.append(cur)
    
    if len(clusters) < 2:
        return None
    
    py_right = max(clusters[0])
    cy_right = max(clusters[1])
    
    cy_lefts = [p[0] for p in num_positions if abs(p[1] - cy_right) < 20]
    py_lefts = [p[0] for p in num_positions if abs(p[1] - py_right) < 20]
    cy_left = min(cy_lefts) if cy_lefts else cy_right - 80
    py_left = min(py_lefts) if py_lefts else py_right - 80
    mid_cy_py = (cy_right + py_left) / 2
    
    note_x0 = note_word['x0'] if note_word else cy_left - 30
    note_right = (note_word['x1'] + 20) if note_word else cy_left - 5
    
    return {
        'particulars_right': note_x0 - 5,
        'note_left': note_x0 - 5,
        'note_right': min(note_right, cy_left - 5),
        'cy_left': cy_left - 10,
        'cy_right': mid_cy_py,
        'py_left': mid_cy_py,
        'py_right': py_right + 20,
    }


def extract_rows_from_page(page, col_bounds=None):
    """Extract structured rows using column boundaries."""
    words = page.extract_words(keep_blank_chars=False, x_tolerance=2, y_tolerance=3)
    if not words:
        return [], col_bounds
    
    if col_bounds is None:
        col_bounds = find_column_boundaries(words, page.width)
    if col_bounds is None:
        return [], None
    
    y_tol = 4
    rows_by_y = defaultdict(list)
    for w in words:
        y_key = round(w['top'] / y_tol) * y_tol
        rows_by_y[y_key].append(w)
    
    result = []
    for y in sorted(rows_by_y.keys()):
        row_words = sorted(rows_by_y[y], key=lambda w: w['x0'])
        
        parts = {'particulars': [], 'note': [], 'cy': [], 'py': []}
        
        for w in row_words:
            txt = w['text'].strip()
            if not txt:
                continue
            x0, x1 = w['x0'], w['x1']
            x_mid = (x0 + x1) / 2
            
            # Assign to column
            if x1 <= col_bounds['particulars_right'] + 5:
                parts['particulars'].append(txt)
            elif x0 >= col_bounds['note_left'] and x1 <= col_bounds['note_right'] + 5:
                if re.match(r'^\d{1,2}(-\d{1,2})?$', txt):
                    parts['note'].append(txt)
                elif x_mid < col_bounds['cy_left']:
                    parts['note'].append(txt)
                else:
                    parts['cy'].append(txt)
            elif x_mid < col_bounds['cy_right'] and x0 >= col_bounds['cy_left'] - 10:
                parts['cy'].append(txt)
            elif x0 >= col_bounds['py_left'] - 10:
                parts['py'].append(txt)
            else:
                # Ambiguous — use proximity
                dist_note = abs(x_mid - (col_bounds['note_left'] + col_bounds['note_right']) / 2)
                dist_cy = abs(x_mid - (col_bounds['cy_left'] + col_bounds['cy_right']) / 2)
                dist_py = abs(x_mid - (col_bounds['py_left'] + col_bounds['py_right']) / 2)
                
                min_dist = min(dist_note, dist_cy, dist_py)
                if min_dist == dist_py:
                    parts['py'].append(txt)
                elif min_dist == dist_cy:
                    parts['cy'].append(txt)
                elif re.match(r'^\d{1,2}$', txt):
                    parts['note'].append(txt)
                else:
                    parts['particulars'].append(txt)
        
        particulars = ' '.join(parts['particulars']).strip()
        note = ' '.join(parts['note']).strip()
        cy_raw = ' '.join(parts['cy']).strip()
        py_raw = ' '.join(parts['py']).strip()
        
        cy_val = parse_value(cy_raw)
        py_val = parse_value(py_raw)
        
        if not particulars and not cy_raw and not py_raw:
            continue
        
        result.append({
            'particulars': particulars,
            'note': note,
            'cy': cy_val,
            'py': py_val,
            'cy_raw': cy_raw,
            'py_raw': py_raw,
        })
    
    return result, col_bounds


def parse_value(s):
    s = s.strip()
    if not s:
        return ''
    if s == '-':
        return '-'
    s_clean = s.replace(' ', '')
    m = re.match(r'^\(([\d,]+\.?\d*)\)$', s_clean)
    if m:
        return -float(m.group(1).replace(',', ''))
    m = re.match(r'^([\d,]+\.?\d*)$', s_clean)
    if m:
        return float(m.group(1).replace(',', ''))
    return s


def is_stop_line(text):
    low = text.lower().strip()
    return any(p in low for p in STOP_PHRASES)


def detect_col_headers(rows):
    for r in rows[:15]:
        for field in ['particulars', 'cy_raw', 'py_raw']:
            text = r.get(field, '')
            dates = re.findall(r'(?:31\s*March\s*\d{4}|March\s*31(?:st)?,?\s*\d{4})', text, re.I)
            if len(dates) >= 2:
                return dates[0].strip(), dates[1].strip()
    # Check combined text
    for r in rows[:15]:
        text = ' '.join([r.get('particulars',''), r.get('cy_raw',''), r.get('py_raw','')])
        dates = re.findall(r'(?:31\s*March\s*\d{4}|March\s*31(?:st)?,?\s*\d{4})', text, re.I)
        if len(dates) >= 2:
            return dates[0].strip(), dates[1].strip()
    return 'Current Year', 'Prior Year'


def filter_data_rows(rows):
    data_starts = ['income', 'revenue', 'assets', 'equity and liabilities',
                   'shareholder', 'expenses', 'non-current', 'non- current',
                   'current assets', 'equity and liabilit']
    
    start = 0
    for i, r in enumerate(rows):
        low = r['particulars'].lower().strip()
        if any(k in low for k in ['particulars', 'all amount', 'all figures', 'statement of',
                                    'balance sheet as at', 'profit and loss', 'profit & loss',
                                    'indian rupees', 'rupees millions', 'lakhs inr',
                                    'for the year', 'as at as at', 'ended ended']):
            start = i + 1
            continue
        if re.match(r'^(as at|for the|ended|note)', low):
            start = i + 1
            continue
        if re.match(r'^\d{1,2}\s+(march|april|may)', low, re.I):
            start = i + 1
            continue
        if any(low.startswith(m) or low == m for m in data_starts):
            start = i
            break
        if isinstance(r['cy'], (int, float)) and r['particulars']:
            start = i
            break
    
    end = len(rows)
    for i in range(start, len(rows)):
        if is_stop_line(rows[i]['particulars']):
            end = i
            break
    
    filtered = []
    for r in rows[start:end]:
        part = r['particulars'].strip()
        if re.match(r'^\d{1,4}$', part) and r['cy'] == '' and r['py'] == '':
            continue
        if re.match(r'^(CIN|ICAI)', part, re.I):
            continue
        filtered.append(r)
    
    return filtered


def detect_company_name(text):
    for line in text.split('\n')[:10]:
        line = line.strip()
        up = line.upper()
        if 'LIMITED' in up and all(k not in up for k in ['REGISTRATION', 'CIN', 'ICAI']):
            for rem in ['CONSOLIDATED STATEMENT OF PROFIT AND LOSS',
                        'STANDALONE STATEMENT OF PROFIT AND LOSS',
                        'CONSOLIDATED BALANCE SHEET', 'STANDALONE BALANCE SHEET',
                        'STATEMENT OF PROFIT AND LOSS', 'STATEMENT OF PROFIT & LOSS',
                        'BALANCE SHEET', 'AS AT']:
                up = up.replace(rem, '').strip()
            name = up.strip(' -')
            if name:
                return name
    for line in text.split('\n')[:5]:
        line = line.strip()
        if line and len(line) > 5 and all(k not in line.upper() for k in
                                           ['BALANCE', 'PROFIT', 'STATEMENT', 'ALL AMOUNT', 'CIN', '(']):
            return line
    return ''


def process_pdf(pdf_path):
    pdf = pdfplumber.open(pdf_path)
    first_text = pdf.pages[0].extract_text() or ''
    company = detect_company_name(first_text)
    
    statements = []
    i = 0
    while i < len(pdf.pages):
        page_text = pdf.pages[i].extract_text() or ''
        stype = classify_page(page_text)
        
        if stype:
            pages = [i]
            j = i + 1
            while j < len(pdf.pages):
                next_text = pdf.pages[j].extract_text() or ''
                next_type = classify_page(next_text)
                if next_type is None and any(kw in next_text.upper() for kw in
                                              ['TOTAL', 'TRADE PAYABLE', 'PROVISIONS',
                                               'TOTAL EQUITY', 'TOTAL ASSETS', 'TOTAL LIABILITIES']):
                    pages.append(j)
                    j += 1
                else:
                    break
            statements.append((stype, pages))
            i = j
        else:
            upper = page_text[:1000].upper()
            if 'BALANCE SHEET' in upper:
                statements.append(('generic_bs', [i]))
            elif any(k in upper for k in ['PROFIT AND LOSS', 'PROFIT & LOSS']):
                statements.append(('generic_pl', [i]))
            i += 1
    
    sheets = []
    for stype, page_indices in statements:
        sheet_name = STMT_NAMES.get(stype, stype)
        all_rows = []
        col_bounds = None
        
        for pi in page_indices:
            page_rows, new_bounds = extract_rows_from_page(pdf.pages[pi], col_bounds)
            if new_bounds:
                col_bounds = new_bounds
            all_rows.extend(page_rows)
        
        if not all_rows:
            continue
        
        col1, col2 = detect_col_headers(all_rows)
        data_rows = filter_data_rows(all_rows)
        
        if data_rows:
            sheets.append((sheet_name, data_rows, col1, col2, company))
    
    pdf.close()
    return sheets


def write_sheet(ws, rows, col1_header, col2_header, company_name='', statement_title=''):
    header_fill = PatternFill('solid', fgColor='2F5496')
    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    title_font = Font(name='Arial', bold=True, size=13)
    sub_font = Font(name='Arial', bold=True, size=10, color='555555')
    data_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    sec_font = Font(name='Arial', size=10, bold=True, color='2F5496')
    total_bdr = Border(top=Side(style='thin', color='2F5496'), bottom=Side(style='double', color='2F5496'))
    sub_bdr = Border(bottom=Side(style='thin', color='AAAAAA'))
    
    ws.column_dimensions['A'].width = 62
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    row = 1
    if company_name:
        ws.cell(row=row, column=1, value=company_name).font = title_font
        row += 1
    if statement_title:
        ws.cell(row=row, column=1, value=statement_title).font = sub_font
        row += 1
    row += 1
    
    for ci, h in enumerate(['Particulars', 'Note', col1_header, col2_header], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center' if ci > 1 else 'left', wrap_text=True)
    row += 1
    
    bold_kw = ['total', 'profit before', 'profit after', 'profit for the',
               'profit/(loss)', 'total income', 'total expense', 'total assets',
               'total equity', 'total liabilities', 'total comprehensive', 'earnings per']
    
    for r in rows:
        part = r['particulars']
        note = r['note']
        cy = r['cy']
        py = r['py']
        
        is_total = any(k in part.lower() for k in bold_kw)
        is_section = (cy == '' and py == '' and part and not re.search(r'\d', str(part)))
        is_subtotal = (not part and (isinstance(cy, (int, float)) or isinstance(py, (int, float))))
        
        font = bold_font if (is_total or is_subtotal) else (sec_font if is_section else data_font)
        
        ws.cell(row=row, column=1, value=part).font = font
        if note:
            c = ws.cell(row=row, column=2, value=note)
            c.font = data_font
            c.alignment = Alignment(horizontal='center')
        
        for vi, val in enumerate([cy, py]):
            col = 3 + vi
            cell = ws.cell(row=row, column=col)
            if isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = '#,##0.00'
            elif val == '-':
                cell.value = '-'
            elif val:
                cell.value = val
            cell.font = font
            cell.alignment = Alignment(horizontal='right')
        
        if is_total:
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = total_bdr
        elif is_subtotal:
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = sub_bdr
        row += 1
    
    ws.freeze_panes = 'A5'


def extract_tables(pdf_path, output_dir=None):
    """Extract BS/PL tables from a PDF and save as Excel."""
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"File not found: {pdf_path}")
        return

    if output_dir is None:
        output_dir = pdf_path.parent
    else:
        output_dir = Path(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\nProcessing: {pdf_path.name}")
    print("-" * 60)
    try:
        sheets = process_pdf(str(pdf_path))
        if sheets:
            wb = Workbook()
            wb.remove(wb.active)
            used = set()
            for name, rows, c1, c2, company in sheets:
                sn = name[:31]
                cnt = 2
                while sn in used:
                    sn = f"{name[:28]}_{cnt}"
                    cnt += 1
                used.add(sn)
                ws = wb.create_sheet(title=sn)
                write_sheet(ws, rows, c1, c2, company, name)
                print(f"  '{sn}': {len(rows)} rows")

            out_name = pdf_path.stem + '_extracted.xlsx'
            out_path = output_dir / out_name
            wb.save(str(out_path))
            print(f"  -> {out_path}")
        else:
            print(f"  No statements found")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback; traceback.print_exc()


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage:")
        print("  python extract_tables.py 1_BS_PL.pdf")
        print("  python extract_tables.py 1_BS_PL.pdf 2_BS_PL.pdf")
        print("  python extract_tables.py 1_BS_PL.pdf --output ./results")
        sys.exit(1)

    out_dir = None
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            out_dir = sys.argv[idx + 1]
        pdf_files = [a for a in sys.argv[1:] if a not in ("--output", out_dir)]
    else:
        pdf_files = sys.argv[1:]

    for f in pdf_files:
        extract_tables(f, out_dir)